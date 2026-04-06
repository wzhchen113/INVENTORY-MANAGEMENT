"""
DoorDash Menu Scraper for 2AM Project
Uses Playwright to navigate the page, take screenshots, and extract menu items.
Generates an Excel sheet with items, guessed ingredients, and amount columns.
"""

import asyncio
import json
import os
import re
import sys
from datetime import datetime

from playwright.async_api import async_playwright
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Config ──────────────────────────────────────────────────────────────────
DOORDASH_URL = "https://www.doordash.com/en/store/2am-project-28984787/"
SCREENSHOT_DIR = os.path.join(os.path.dirname(__file__), "screenshots")
OUTPUT_DIR = os.path.dirname(os.path.dirname(__file__))  # project root
OUTPUT_FILE = os.path.join(OUTPUT_DIR, "2AM_Project_Menu_Ingredients.xlsx")


# ── Ingredient Guesses ──────────────────────────────────────────────────────
# Maps menu item keywords to likely ingredients
INGREDIENT_MAP = {
    "Chicken Quesadilla": ["Flour Tortilla", "Chicken Breast", "Shredded Cheese", "Seasoning Blend", "Oil"],
    "Mozzarella Sticks": ["Mozzarella Cheese", "Breadcrumbs", "Flour", "Eggs", "Italian Seasoning", "Oil"],
    "Cajun Fries": ["Potatoes", "Cajun Seasoning", "Oil", "Salt"],
    "French Fries": ["Potatoes", "Oil", "Salt"],
    "Onion Rings": ["Onions", "Flour", "Guinness Stout Beer", "Whiskey Flavoring", "Breadcrumbs", "Oil", "Salt"],
    "Cheesesteak Quesadilla": ["Flour Tortilla", "Sliced Steak", "American Cheese", "Bell Peppers", "Onions", "Oil"],
    "Mac N' Cheese Bites": ["Elbow Macaroni", "Cheddar Cheese Sauce", "Breadcrumbs", "Flour", "Eggs", "Oil"],
    "2AM Nuggets": ["Chicken Breast", "Flour", "Breadcrumbs", "Eggs", "Seasoning Blend", "Oil"],
    "2AM Hashbrown": ["Shredded Potatoes", "Oil", "Salt", "Pepper"],
    "Crabmeat Fries": ["Potatoes", "Crabmeat", "Cajun Seasoning", "2AM House Sauce", "Oil", "Salt"],
    "Hush Puppies": ["Cornmeal", "Flour", "Eggs", "Buttermilk", "Onions", "Sugar", "Baking Powder", "Oil"],
    "Sweet Potato Fries": ["Sweet Potatoes", "Oil", "Salt", "Seasoning Blend"],
    "Mac N Cheese": ["Elbow Macaroni", "White Cheddar Cheese", "Milk", "Butter", "Flour", "Salt", "Pepper"],
    "Build-A-Catch": ["Shrimp", "Snow Crab Legs", "Crawfish", "Mussels", "Corn on the Cob", "Potatoes", "Cajun Seasoning", "Garlic Butter Sauce", "Lemon Pepper Sauce", "Old Bay Seasoning"],
    "Philly Cheesesteak": ["Sub Roll", "Sliced Steak", "American Cheese", "Bell Peppers", "Onions", "Mushrooms", "Oil"],
    "Shrimp Po Boy": ["Sub Roll", "Shrimp", "Flour", "Cornmeal", "Lettuce", "Tomatoes", "2AM Sauce", "Oil"],
    "Fried Fish Sandwich": ["Sub Roll", "Whiting Fish", "Flour", "Cornmeal", "Tartar Sauce", "Lettuce", "Oil"],
    "Chicken Sandwich": ["Brioche Bun", "Chicken Breast", "Flour", "Pickles", "Mayo", "Lettuce", "Oil"],
    "Spicy Chicken Sandwich": ["Brioche Bun", "Chicken Breast", "Flour", "Spicy Seasoning", "2AM Sauce", "Pickles", "Lettuce", "Oil"],
    "Chicken Cheesesteak": ["Sub Roll", "Grilled Chicken", "American Cheese", "Bell Peppers", "Onions", "Oil"],
    "6 Wings": ["Chicken Wings", "Flour", "Seasoning Blend", "Wing Sauce", "Oil"],
    "10 Wings": ["Chicken Wings", "Flour", "Seasoning Blend", "Wing Sauce", "Oil"],
    "20 Wings": ["Chicken Wings", "Flour", "Seasoning Blend", "Wing Sauce", "Oil"],
    "Combo Over Rice": ["Lamb Meat", "Chicken Breast", "Yellow Rice", "Yogurt Sauce", "Lettuce", "Tomatoes", "Seasoning Blend", "Oil"],
    "Chicken Over Rice": ["Chicken Breast", "Yellow Rice", "Yogurt Sauce", "Lettuce", "Tomatoes", "Seasoning Blend", "Oil"],
    "Lamb Over Rice": ["Lamb Meat", "Yellow Rice", "Yogurt Sauce", "Lettuce", "Tomatoes", "Seasoning Blend", "Oil"],
    "Shrimp Over Rice": ["Shrimp", "Yellow Rice", "Yogurt Sauce", "Lettuce", "Tomatoes", "Seasoning Blend", "Oil"],
    "Steak Over Rice": ["Sliced Steak", "Yellow Rice", "Yogurt Sauce", "Lettuce", "Tomatoes", "Seasoning Blend", "Oil"],
    "Lamb Gyro": ["Pita Bread", "Lamb Meat", "Lettuce", "Tomatoes", "Onions", "White Sauce", "Seasoning Blend"],
    "Chicken Gyro": ["Pita Bread", "Chicken Breast", "Lettuce", "Tomatoes", "Onions", "White Sauce", "Seasoning Blend"],
    "Combo Gyro": ["Pita Bread", "Lamb Meat", "Chicken Breast", "Lettuce", "Tomatoes", "Onions", "White Sauce", "Seasoning Blend"],
    "Steak Gyro": ["Pita Bread", "Sliced Steak", "Lettuce", "Tomatoes", "Onions", "White Sauce", "Seasoning Blend"],
    "Shrimp Gyro": ["Pita Bread", "Shrimp", "Lettuce", "Tomatoes", "Onions", "White Sauce", "Seasoning Blend"],
    "Bird & Buried": ["Chicken Tenders", "Flour", "Breadcrumbs", "Eggs", "Potatoes", "Oil", "Salt", "Seasoning Blend"],
    "Fish & Chips": ["Whiting Fish", "Flour", "Cornmeal", "Potatoes", "Oil", "Salt", "Tartar Sauce"],
    "Surf & Turf": ["Shrimp", "Flour", "Cornmeal", "Potatoes", "Oil", "Salt", "Cocktail Sauce"],
    "Choco Cake Slice": ["Chocolate Cake (pre-made)", "Chocolate Mousse Filling", "Fudge Icing"],
    "Crème Brulee": ["Crème Fraîche", "Whole Milk", "Eggs", "Sugar", "Vanilla Extract"],
    "Chocolate Mousse": ["Belgian Chocolate", "Heavy Cream", "Eggs", "Sugar", "Cookie Crumble", "Chocolate Sauce"],
    "Tres Leches": ["Sponge Cake", "Evaporated Milk", "Sweetened Condensed Milk", "Heavy Cream", "Whipped Cream"],
    "Cheesecake Slice": ["Cheesecake (pre-made)", "Graham Cracker Crust", "Vanilla Bean"],
    "Banana Pudding": ["Graham Cracker Crumbs", "Bananas", "Banana Crème Filling", "Mousse Topping"],
    "Caramel Mousse": ["Heavy Cream", "Sugar", "Caramel Sauce", "Sea Salt", "Gelatin"],
    "Triple Mix": ["Specialty Drink Mix (house blend)", "Ice"],
    "Half N Half Tea": ["Lemonade", "Iced Tea", "Sugar", "Ice"],
    "Thai Iced Tea": ["Black Tea", "Sweetened Condensed Milk", "Evaporated Milk", "Sugar", "Ice"],
    "Can Soda": ["Assorted Canned Sodas"],
    "Bottled Water": ["Bottled Water"],
    "2AM Cheeseburger": ["Burger Bun", "Ground Beef Patty", "American Cheese", "Lettuce", "Tomatoes", "Onions", "Pickles", "Ketchup", "Mustard", "Oil"],
    "Blue Cheese": ["Blue Cheese Sauce (pre-made)"],
    "Honey Mustard": ["Honey Mustard Sauce (pre-made)"],
    "BBQ Sauce": ["BBQ Sauce (pre-made)"],
    "Marinara Sauce": ["Marinara Sauce (pre-made)"],
    "Sour Cream": ["Sour Cream"],
    "2AM Sauce": ["2AM House Sauce (proprietary)"],
    "Ranch": ["Ranch Dressing (pre-made)"],
    "White Sauce": ["White Sauce (yogurt-based, house recipe)"],
}

# Full menu data (from Uber Eats verified source)
MENU_ITEMS = [
    # Appetizers
    ("Appetizers", "Chicken Quesadilla", 12.99, "Late night favorite. Seasoned Chicken + Cheese = Quesadilla."),
    ("Appetizers", "Mozzarella Sticks", 7.49, "Stretchy warm cheese with a crunchy exterior."),
    ("Appetizers", "Cajun Fries", 7.29, "Fries tossed in cajun seasoning."),
    ("Appetizers", "French Fries", 6.49, "Classic crispy french fries."),
    ("Appetizers", "Onion Rings", 7.49, "Battered with Guinness stout beer and whiskey flavoring."),
    ("Appetizers", "Cheesesteak Quesadilla", 14.99, "Tender steak and melted cheese in a soft tortilla."),
    ("Appetizers", "Mac N' Cheese Bites", 8.49, "Deep fried bites of gooey mac and cheese."),
    ("Appetizers", "2AM Nuggets", 8.99, "Who doesn't want chicken nuggets at 2AM?"),
    ("Appetizers", "2AM Hashbrown", 2.49, "Fluffy inside, crispy and toasty outside."),
    ("Appetizers", "Crabmeat Fries", 12.99, "2AM Project Bestseller. Cajun fries topped with crabmeat and 2AM sauce."),
    ("Appetizers", "Hush Puppies", 7.49, "Baltimorean appetizer, popular with locals."),
    ("Appetizers", "Sweet Potato Fries", 6.49, "Seasoned sweet potato fries fried to perfection."),
    ("Appetizers", "Mac N Cheese", 8.49, "Creamy late night macaroni & cheese."),
    # Seafood Boils
    ("Seafood Boils", "Build-A-Catch", 0.00, "Customize your boil bag! Pick at least 2 seafoods, choose sauce and spice level. Corn & potatoes included."),
    # Sandwiches
    ("Sandwiches", "Philly Cheesesteak", 15.49, "2AM PROJECT Bestseller. Thinly sliced beef with sautéed veggies."),
    ("Sandwiches", "Shrimp Po Boy", 15.49, "Fried shrimp on a sub with lettuce and tomatoes, drizzled with 2AM sauce."),
    ("Sandwiches", "Fried Fish Sandwich", 15.49, "Fried whiting on a sub with tartar sauce."),
    ("Sandwiches", "Chicken Sandwich", 14.99, "Fried chicken sandwich with mayo and lettuce."),
    ("Sandwiches", "Spicy Chicken Sandwich", 14.99, "Bold flavors with signature 2AM sauce."),
    ("Sandwiches", "Chicken Cheesesteak", 14.99, "Grilled chicken topped with cheese and green peppers on a fresh sub."),
    # Burgers
    ("Burgers", "2AM Cheeseburger", 13.99, "Smash burger with cheese, lettuce, tomato, onion, pickles."),
    # Wings
    ("Wings", "6 Wings", 11.99, "Crispy golden wings tossed in your choice of sauce."),
    ("Wings", "10 Wings", 17.99, "Crispy golden wings tossed in your choice of sauce."),
    ("Wings", "20 Wings", 32.99, "Crispy golden wings tossed in your choice of sauce."),
    # Rice Platters
    ("Rice Platters", "Combo Over Rice", 14.49, "Lamb and grilled chicken over yellow rice with yogurt sauce, lettuce, tomatoes."),
    ("Rice Platters", "Chicken Over Rice", 13.99, "Grilled chicken over yellow rice with yogurt sauce, lettuce, tomatoes."),
    ("Rice Platters", "Lamb Over Rice", 13.99, "Lamb over yellow rice with yogurt sauce, lettuce, tomatoes."),
    ("Rice Platters", "Shrimp Over Rice", 17.99, "Grilled shrimp over yellow rice with yogurt sauce, lettuce, tomatoes."),
    ("Rice Platters", "Steak Over Rice", 14.49, "Grilled steak over yellow rice with yogurt sauce, lettuce, tomatoes."),
    # Gyros
    ("Gyros", "Lamb Gyro", 11.99, "Grilled lamb with lettuce, tomato, onion, white sauce in pita."),
    ("Gyros", "Chicken Gyro", 11.99, "Grilled chicken with lettuce, tomato, onion, white sauce in pita."),
    ("Gyros", "Combo Gyro", 12.49, "Lamb & chicken with lettuce, tomato, onion, white sauce in pita."),
    ("Gyros", "Steak Gyro", 11.99, "Grilled steak with lettuce, tomato, onion, white sauce in pita."),
    ("Gyros", "Shrimp Gyro", 15.99, "Grilled shrimp with lettuce, tomato, onion, white sauce in pita."),
    # The Duos
    ("The Duos", "Bird & Buried", 16.49, "Crispy chicken tenders fried to perfection, served over fries."),
    ("The Duos", "Fish & Chips", 18.49, "Fried whiting fish served over fresh fries."),
    ("The Duos", "Surf & Turf", 18.49, "Fried shrimp served over fresh fries."),
    # Desserts
    ("Desserts", "Choco Cake Slice", 8.99, "Two-layer chocolate cake with chocolate mousse filling, covered in fudge icing."),
    ("Desserts", "Crème Brulee", 7.49, "Farm-fresh crème fraîche, whole milk, eggs and sugar."),
    ("Desserts", "Chocolate Mousse", 7.49, "Belgian chocolate mousse on cookie crumble, dipped in chocolate sauce."),
    ("Desserts", "Tres Leches", 7.49, "Sponge cake soaked with evaporated milk, condensed milk, and heavy cream."),
    ("Desserts", "Cheesecake Slice", 8.99, "Graham cracker crust, rich cheesecake filling with vanilla bean speckles."),
    ("Desserts", "Banana Pudding", 7.49, "Parfait with graham cracker crumb, banana crème filling, and mousse."),
    ("Desserts", "Caramel Mousse", 7.49, "Mousse layered with caramel swirls, topped with caramel drizzle and sea salt."),
    # Beverages
    ("Beverages", "Triple Mix", 5.49, "2AM Project's specialty drink (comes in different colors every time)."),
    ("Beverages", "Half N Half Tea", 5.49, "50% lemonade and 50% iced tea."),
    ("Beverages", "Thai Iced Tea", 5.49, "Thai iced tea with condensed milk."),
    ("Beverages", "Can Soda", 2.49, "A can of soda."),
    ("Beverages", "Bottled Water", 1.49, "A bottle of water."),
    # Signature Sauces
    ("Signature Sauces", "Blue Cheese", 0.95, ""),
    ("Signature Sauces", "Honey Mustard", 0.95, ""),
    ("Signature Sauces", "BBQ Sauce", 0.95, ""),
    ("Signature Sauces", "Marinara Sauce", 0.95, ""),
    ("Signature Sauces", "Sour Cream", 0.95, ""),
    ("Signature Sauces", "2AM Sauce", 0.95, ""),
    ("Signature Sauces", "Ranch", 0.95, ""),
    ("Signature Sauces", "White Sauce", 0.95, ""),
]


async def scrape_doordash():
    """Navigate to DoorDash, scroll through the page, take screenshots, and extract menu data."""
    os.makedirs(SCREENSHOT_DIR, exist_ok=True)

    print("[1/4] Launching browser...")
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        context = await browser.new_context(
            viewport={"width": 1400, "height": 900},
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/125.0.0.0 Safari/537.36"
            ),
        )
        page = await context.new_page()

        print(f"[2/4] Navigating to DoorDash...")
        try:
            await page.goto(DOORDASH_URL, wait_until="domcontentloaded", timeout=30000)
        except Exception as e:
            print(f"  Warning: page load issue ({e}), continuing anyway...")

        # Wait for content to render
        await page.wait_for_timeout(5000)

        # Take initial screenshot
        await page.screenshot(path=os.path.join(SCREENSHOT_DIR, "01_initial_load.png"), full_page=False)
        print("  Screenshot: 01_initial_load.png")

        # Close any popups/modals
        for selector in [
            'button[aria-label="Close"]',
            'button:has-text("Close")',
            '[data-testid="close-button"]',
            'button:has-text("Not now")',
        ]:
            try:
                btn = page.locator(selector).first
                if await btn.is_visible(timeout=1000):
                    await btn.click()
                    await page.wait_for_timeout(500)
            except:
                pass

        print("[3/4] Scrolling and capturing menu sections...")
        scraped_items = []
        screenshot_count = 2

        # Scroll down the page in increments, taking screenshots and extracting text
        scroll_position = 0
        page_height = await page.evaluate("document.body.scrollHeight")
        viewport_height = 900

        while scroll_position < page_height:
            await page.evaluate(f"window.scrollTo(0, {scroll_position})")
            await page.wait_for_timeout(1500)

            # Take screenshot every few scrolls
            if screenshot_count <= 10:
                fname = f"{screenshot_count:02d}_scroll_{scroll_position}.png"
                await page.screenshot(path=os.path.join(SCREENSHOT_DIR, fname), full_page=False)
                print(f"  Screenshot: {fname}")
                screenshot_count += 1

            # Try to extract menu item elements from the visible DOM
            items = await page.evaluate("""
                () => {
                    const results = [];

                    // DoorDash uses various selectors for menu items
                    // Try common patterns
                    const selectors = [
                        '[data-testid="StoreMenuItem"]',
                        '[data-testid="MenuItem"]',
                        '.sc-bczRLJ',  // styled component menu items
                        'button[kind="ITEM"]',
                        'div[class*="MenuItem"]',
                        'div[class*="menu-item"]',
                        'a[href*="/item/"]',
                    ];

                    for (const sel of selectors) {
                        const elements = document.querySelectorAll(sel);
                        for (const el of elements) {
                            // Look for item name and price within the element
                            const nameEl = el.querySelector('span, h3, h4, [class*="name"], [class*="Name"], [class*="title"], [class*="Title"]');
                            const priceEl = el.querySelector('[class*="price"], [class*="Price"], [class*="cost"]');
                            const descEl = el.querySelector('[class*="desc"], [class*="Desc"], [class*="description"]');

                            if (nameEl) {
                                const name = nameEl.textContent.trim();
                                const price = priceEl ? priceEl.textContent.trim() : '';
                                const desc = descEl ? descEl.textContent.trim() : '';
                                if (name && name.length > 1 && name.length < 100) {
                                    results.push({name, price, desc});
                                }
                            }
                        }
                    }

                    // Also try extracting from any visible text that looks like menu items
                    const allSpans = document.querySelectorAll('span, h3, h4, p, div');
                    for (const el of allSpans) {
                        const text = el.textContent.trim();
                        // Look for price patterns like $X.XX next to item names
                        const priceMatch = text.match(/\\$\\d+\\.\\d{2}/);
                        if (priceMatch && text.length < 200) {
                            const cleanText = text.replace(/\\$\\d+\\.\\d{2}/, '').trim();
                            if (cleanText.length > 2 && cleanText.length < 80) {
                                results.push({
                                    name: cleanText,
                                    price: priceMatch[0],
                                    desc: ''
                                });
                            }
                        }
                    }

                    return results;
                }
            """)
            scraped_items.extend(items)

            scroll_position += viewport_height - 100
            # Re-check page height (lazy loading may have added content)
            page_height = await page.evaluate("document.body.scrollHeight")

        # Take a full-page screenshot at the end
        try:
            await page.screenshot(path=os.path.join(SCREENSHOT_DIR, "full_page.png"), full_page=True)
            print("  Screenshot: full_page.png")
        except:
            print("  (full-page screenshot too large, skipped)")

        # Deduplicate scraped items
        seen = set()
        unique_scraped = []
        for item in scraped_items:
            key = item["name"].lower().strip()
            if key not in seen:
                seen.add(key)
                unique_scraped.append(item)

        print(f"\n  Scraped {len(unique_scraped)} unique items from DoorDash DOM")
        if unique_scraped:
            print("  Sample items found:")
            for item in unique_scraped[:10]:
                print(f"    - {item['name']} {item['price']}")

        await browser.close()

    return unique_scraped


def generate_excel(scraped_items):
    """Generate the Excel sheet with menu items, ingredients, and amount columns."""
    print("\n[4/4] Generating Excel spreadsheet...")

    wb = Workbook()

    # ── Styles ───────────────────────────────────────────────────────────
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    category_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    category_fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
    ingredient_header_fill = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")
    amount_header_fill = PatternFill(start_color="F39C12", end_color="F39C12", fill_type="solid")
    item_font = Font(name="Calibri", size=11)
    price_font = Font(name="Calibri", size=11, color="27AE60", bold=True)
    light_gray_fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin", color="DEE2E6"),
        right=Side(style="thin", color="DEE2E6"),
        top=Side(style="thin", color="DEE2E6"),
        bottom=Side(style="thin", color="DEE2E6"),
    )

    # ── Sheet 1: Menu Items with Ingredients ─────────────────────────────
    ws = wb.active
    ws.title = "Menu & Ingredients"

    # Find max ingredients count for column sizing
    max_ingredients = max(len(v) for v in INGREDIENT_MAP.values())

    # Headers
    headers = ["Category", "Menu Item", "Price", "Description"]
    for i in range(1, max_ingredients + 1):
        headers.append(f"Ingredient {i}")
        headers.append(f"Amount {i}")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
        if "Amount" in header:
            cell.fill = amount_header_fill
        elif "Ingredient" in header:
            cell.fill = ingredient_header_fill
        else:
            cell.fill = header_fill

    # Column widths
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 55
    for i in range(5, len(headers) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 18

    # Data rows
    row = 2
    current_category = None
    for category, name, price, desc in MENU_ITEMS:
        # Category separator row
        if category != current_category:
            current_category = category
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=row, column=col)
                cell.fill = category_fill
                cell.font = category_font
                cell.border = thin_border
            ws.cell(row=row, column=1, value=category).font = category_font
            ws.cell(row=row, column=1).fill = category_fill
            row += 1

        # Item row
        fill = light_gray_fill if (row % 2 == 0) else PatternFill()

        ws.cell(row=row, column=1, value=category).font = item_font
        ws.cell(row=row, column=1).fill = fill
        ws.cell(row=row, column=1).border = thin_border

        ws.cell(row=row, column=2, value=name).font = Font(name="Calibri", size=11, bold=True)
        ws.cell(row=row, column=2).fill = fill
        ws.cell(row=row, column=2).border = thin_border

        price_str = f"${price:.2f}" if price > 0 else "Varies"
        ws.cell(row=row, column=3, value=price_str).font = price_font
        ws.cell(row=row, column=3).fill = fill
        ws.cell(row=row, column=3).border = thin_border
        ws.cell(row=row, column=3).alignment = Alignment(horizontal="center")

        ws.cell(row=row, column=4, value=desc).font = item_font
        ws.cell(row=row, column=4).fill = fill
        ws.cell(row=row, column=4).border = thin_border
        ws.cell(row=row, column=4).alignment = Alignment(wrap_text=True)

        # Ingredients and Amount columns
        ingredients = INGREDIENT_MAP.get(name, [])
        col = 5
        for ingredient in ingredients:
            # Ingredient name
            cell = ws.cell(row=row, column=col, value=ingredient)
            cell.font = item_font
            cell.fill = fill
            cell.border = thin_border
            col += 1
            # Amount (empty - user fills this in)
            cell = ws.cell(row=row, column=col)
            cell.fill = PatternFill(start_color="FFF9E6", end_color="FFF9E6", fill_type="solid")
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")
            cell.number_format = '0.00'
            col += 1

        # Fill remaining ingredient/amount columns with border
        while col <= len(headers):
            cell = ws.cell(row=row, column=col)
            cell.fill = fill
            cell.border = thin_border
            col += 1

        row += 1

    # Freeze header row
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{row - 1}"

    # ── Sheet 2: Ingredient Master List ──────────────────────────────────
    ws2 = wb.create_sheet("Ingredient Master List")

    # Collect all unique ingredients
    all_ingredients = set()
    for ingredients in INGREDIENT_MAP.values():
        all_ingredients.update(ingredients)
    all_ingredients = sorted(all_ingredients)

    headers2 = ["Ingredient", "Unit", "Cost Per Unit", "Supplier", "Notes"]
    for col, header in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    ws2.column_dimensions["A"].width = 35
    ws2.column_dimensions["B"].width = 12
    ws2.column_dimensions["C"].width = 15
    ws2.column_dimensions["D"].width = 25
    ws2.column_dimensions["E"].width = 30

    for i, ingredient in enumerate(all_ingredients, 2):
        fill = light_gray_fill if (i % 2 == 0) else PatternFill()
        ws2.cell(row=i, column=1, value=ingredient).font = item_font
        ws2.cell(row=i, column=1).fill = fill
        ws2.cell(row=i, column=1).border = thin_border
        for col in range(2, 6):
            cell = ws2.cell(row=i, column=col)
            cell.fill = PatternFill(start_color="FFF9E6", end_color="FFF9E6", fill_type="solid")
            cell.border = thin_border
            if col == 3:
                cell.number_format = '$#,##0.00'

    ws2.freeze_panes = "A2"

    # ── Sheet 3: Scraped Raw Data (from Playwright) ──────────────────────
    ws3 = wb.create_sheet("Scraped Raw Data")
    headers3 = ["Item Name (from DoorDash)", "Price", "Description"]
    for col, header in enumerate(headers3, 1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = PatternFill(start_color="8E44AD", end_color="8E44AD", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    ws3.column_dimensions["A"].width = 40
    ws3.column_dimensions["B"].width = 12
    ws3.column_dimensions["C"].width = 60

    if scraped_items:
        for i, item in enumerate(scraped_items, 2):
            ws3.cell(row=i, column=1, value=item["name"]).border = thin_border
            ws3.cell(row=i, column=2, value=item["price"]).border = thin_border
            ws3.cell(row=i, column=3, value=item["desc"]).border = thin_border
    else:
        ws3.cell(row=2, column=1, value="(DoorDash blocked scraping — verified data is in Sheet 1)")

    ws3.freeze_panes = "A2"

    # ── Save ─────────────────────────────────────────────────────────────
    wb.save(OUTPUT_FILE)
    print(f"\n[OK] Excel saved to: {OUTPUT_FILE}")
    print(f"  Screenshots saved to: {SCREENSHOT_DIR}/")
    print(f"\n  Sheet 1: 'Menu & Ingredients' - {len(MENU_ITEMS)} items with guessed ingredients")
    print(f"           Yellow cells = enter your amounts here")
    print(f"  Sheet 2: 'Ingredient Master List' - {len(all_ingredients)} unique ingredients")
    print(f"           Fill in unit, cost, supplier info")
    print(f"  Sheet 3: 'Scraped Raw Data' - raw items from Playwright scrape")


async def main():
    print("=" * 60)
    print("  2AM Project — DoorDash Menu Scraper")
    print("=" * 60)
    scraped = await scrape_doordash()
    generate_excel(scraped)
    print("\nDone!")


if __name__ == "__main__":
    asyncio.run(main())
